import os
from string import ascii_uppercase as alphabet
from tqdm import tqdm
import win32com.client as win32
from openpyxl import load_workbook


def exchange(fold):
    """
    :param fold: 文件夹
    :return:
    """
    path = os.getcwd()
    path = os.path.join(path, fold)
    files = os.listdir(path)
    for file_name in files:
        if file_name.rsplit('.', 1)[-1] == 'xls':
            fname = os.path.join(path, file_name)
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(fname)
            # 在原来的位置创建出：原名+'.xlsx'文件
            wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            wb.Close()  # FileFormat = 56 is for .xls extension
            excel.Application.Quit()
            os.remove(fname)
            return file_name.rsplit('.', 1)[0] + '.xlsx'
        if file_name.rsplit('.', 1)[-1] == 'xlsx':
            return file_name


class XLSX:
    def __init__(self, path: str, headList: list, sheetNum=0):
        self.path = path
        self.sheetNum = sheetNum
        self.wb = load_workbook(path)
        self.workSheet = self.wb[self.wb.sheetnames[sheetNum]]
        self.head = headList  # 新添加的表头
        self.addHead(self.head)
        # # 填充相同值
        # self.result = self.getAllData()
        # for key in self.result.keys():
        #     self.writeUrlAndData(key, self.result[key])
        # 要进行爬虫的url队列
        urlDict = {}
        print("初始化任务队列：  需要首先遍历Excel...")
        for num in tqdm(range(2, self.workSheet.max_row + 1)):
            if self.readXlsx(row=num, col=self.workSheet.max_column - len(self.head) + 1) == None:
                urlDict[self.readXlsx(row=num, col=1)] = None
        self.urlList = list(urlDict.keys())

    def writeXlsx(self, row: int, col: int, value: str):
        """
        写入表格，输入行和列，起始位置为第一行第一列\n
        :param row:
        :param col:
        :param value:
        :return:
        """
        colKey = alphabet[col - 1]
        rowKey = row
        key = "{}{}".format(colKey, rowKey)
        self.workSheet[key] = value

    def readXlsx(self, row: int, col: int):
        """
        读取表格，输入行和列，起始位置为第一行第一列\n
        :param row:
        :param col:
        :return:
        """
        colKey = alphabet[col - 1]
        rowKey = row
        key = "{}{}".format(colKey, rowKey)
        return self.workSheet[key].value

    def addHead(self, headList: list):
        """
        在末尾添加表头，在最后几个表头不是输入参数的情况下
        :param headList:
        :return:
        """
        self.head = headList
        flag = True
        maxCol = self.workSheet.max_column
        length = len(headList)
        if maxCol > length:
            for i in range(length):
                temp = self.readXlsx(row=1, col=maxCol - length + i + 1)
                if temp != headList[i]:
                    flag = False
                    break
        else:
            flag = False
        if flag == False:
            for i in range(length):
                self.writeXlsx(row=1, col=maxCol + i + 1, value=headList[i])
        return

    def getAllData(self):
        """
        获取xlsx中所有已经填写的数据，便于查重\n
        :return:
        """
        result = {}
        maxCol = self.workSheet.max_column
        maxRow = self.workSheet.max_row
        for i in range(2, maxRow + 1):
            if self.readXlsx(row=i, col=maxCol - len(self.head) + 1) == '是':
                result[self.readXlsx(row=i, col=1)] = []
                for num in range(len(self.head)):
                    result[self.readXlsx(row=i, col=1)].append(
                        self.readXlsx(row=i, col=maxCol - len(self.head) + num + 1))
        return result

    def save(self):
        """
        保存\n
        :return:
        """
        try:
            self.wb.save(self.path)
        except:
            a = None

    def getUrl(self):
        """
        获取一条还没有进行爬虫操作或者是本次还没有进行重复爬虫的一条失败链接\n
        :return: 链接url
        """
        return self.urlList.pop(0)

    def writeUrlAndData(self, url, data):
        """
        写入一条数据到xlsx中，并且填充重复内容\n
        :param url:
        :param data:
        :return:
        """
        # 找到此url所在的行
        maxRow = self.workSheet.max_row
        maxCol = self.workSheet.max_column
        rowList = []
        for row in range(2, maxRow + 1):
            if self.readXlsx(row=row, col=1) == url and self.readXlsx(row=row, col=maxCol - len(data) + 1) != "是":
                rowList.append(row)
        # 每一行url相同就进行写入
        for row in rowList:
            for num in range(len(self.head)):
                self.writeXlsx(row=row, col=maxCol - len(data) + 1 + num, value=data[num])
        # 保存操作
        self.save()


if __name__ == '__main__':
    basePath = r'./data'
    xlsxName = exchange('./data')
    xlsxPath = os.path.join(basePath, xlsxName)
    xlsx = XLSX(xlsxPath, ["是否完成爬虫", "店铺名", "PDF路径"])
    a = 1
