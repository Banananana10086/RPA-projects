import time
import configparser

import win32com.client as win32
import os
import re
from openpyxl import load_workbook
import openpyxl
from string import ascii_uppercase as alphabet


def exchange(fold):
    """
    :param fold: 文件夹
    :return:
    """
    path = os.getcwd()
    path = os.path.join(path, fold)
    files = os.listdir(path)
    for file_name in files:
        if file_name.rsplit('.', 1)[-1] == 'xls':
            fname = os.path.join(path, file_name)
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(fname)
            # 在原来的位置创建出：原名+'.xlsx'文件
            wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            wb.Close()  # FileFormat = 56 is for .xls extension
            excel.Application.Quit()
            os.remove(fname)
            return file_name.rsplit('.', 1)[0] + '.xlsx'
        if file_name.rsplit('.', 1)[-1] == 'xlsx':
            return file_name


def writeXlsx(workSheet, row, col, value):
    """
    写入表格，输入行和列，起始位置为第一行第一列\n
    :param workSheet: 列表对象
    :param row: 行
    :param col: 列
    :param value: 值
    :return:
    """
    colKey = alphabet[col - 1]
    rowKey = row
    key = "{}{}".format(colKey, rowKey)
    workSheet[key] = value


def readXlsx(workSheet, row, col):
    """
    读取表格，输入行和列，起始位置为第一行第一列
    :param workSheet: 列表对象
    :param row: 行
    :param col: 列
    :return: value
    """
    colKey = alphabet[col - 1]
    rowKey = row
    key = "{}{}".format(colKey, rowKey)
    return workSheet[key].value


def addHead(workSheet, headList: list):
    """
    在末尾添加表头，在最后几个表头不是输入参数的情况下
    :param workSheet:
    :param headList:
    :return:
    """
    flag = True
    maxCol = workSheet.max_column
    length = len(headList)
    if maxCol > length:
        for i in range(length):
            temp = readXlsx(workSheet, row=1, col=maxCol - length + i + 1)
            if temp != headList[i]:
                flag = False
                break
    else:
        flag = False
    if flag == False:
        for i in range(length):
            writeXlsx(workSheet, row=1, col=maxCol + i + 1, value=headList[i])
    return


def getAllData(workSheet):
    """
    获取xlsx中所有已经填写的数据，便于查重\n
    :param workSheet:
    :return:
    """
    result = {}
    maxCol = workSheet.max_column
    maxRow = workSheet.max_row
    for i in range(2, maxRow + 1):
        if readXlsx(workSheet, row=i, col=maxCol - 2) == '是':
            result[readXlsx(workSheet, row=i, col=1)] = [readXlsx(workSheet, row=i, col=maxCol - 2),
                                                         readXlsx(workSheet, row=i, col=maxCol - 1),
                                                         readXlsx(workSheet, row=i, col=maxCol)]
    return result


if __name__ == '__main__':
    xlsxName = exchange('data')
    xlsxPath = r'./data/{}'.format(xlsxName)
    wb = load_workbook(xlsxPath)
    sheet = wb[wb.sheetnames[0]]
    headList = ['是否已经获取数据', '店铺名', 'PDF路径']
    addHead(sheet, headList)
    wb.save(xlsxPath)
